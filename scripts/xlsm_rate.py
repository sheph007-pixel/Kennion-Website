#!/usr/bin/env python3
"""Server-side rate engine that delegates to the actuary's xlsm.

Pipeline:
  1. Open xlsm via LibreOffice (python-uno if available, else CLI)
  2. Inject Inputs + Census
  3. Force full recalc
  4. Read plan_rates back from '3. Rate Sheet - HDV'

Two execution paths, selected at runtime based on what's installed:

  path A (UNO):      soffice --accept=socket + python-uno
                     - opens xlsm directly (VBA + named ranges preserved)
                     - writes cells via UNO API in bulk (setDataArray)
                     - doc.calculateAll() forces real recalc
                     - reads rates via UNO
                     - FASTEST + MOST ROBUST when python3-uno is installed

  path B (CLI fallback): openpyxl + soffice --convert-to xlsx
                     - less robust (openpyxl round-trip can drop defined names)
                     - but works even without python-uno

Path selection is logged loudly to stderr so we can see in Railway logs.
"""
from __future__ import annotations
import json, os, shutil, subprocess, sys, tempfile, time, uuid, traceback
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

CENSUS_FIRST_ROW_1 = 18          # 1-indexed
CENSUS_FIRST_ROW_0 = 17          # 0-indexed
CENSUS_MAX_ROWS = 300
# 0-indexed (col, row) for Inputs
INP_B4_0 = (1, 3)
INP_B5_0 = (1, 4)
INP_B7_0 = (1, 6)
INP_E7_0 = (4, 6)
# Census cols 0-indexed
CEN_COL_SSN, CEN_COL_REL, CEN_COL_FN, CEN_COL_LN = 2, 3, 4, 5
CEN_COL_DOB, CEN_COL_COMPANY, CEN_COL_PLAN, CEN_COL_COV, CEN_COL_COST = 6, 7, 8, 9, 10
# HDV plan rows 0-indexed (1-indexed 11..22)
HDV_PLAN_ROWS_0 = list(range(10, 22))


def _log(msg: str) -> None:
    sys.stderr.write(f"[xlsm_rate] {time.time():.1f} {msg}\n"); sys.stderr.flush()


def _fail(msg: str, extra: Optional[dict] = None) -> None:
    out = {"error": msg}
    if extra: out.update(extra)
    sys.stdout.write(json.dumps(out) + "\n"); sys.exit(0)


def _norm_rel(rel: str) -> str:
    r = (rel or "").strip().lower()
    if r in ("ee","employee","emp","subscriber","self","primary"): return "Employee"
    if r in ("sp","spouse","partner") or "partner" in r: return "Spouse"
    return "Child"


def _parse_dob(s: str) -> datetime:
    s = (s or "").strip()
    for fmt in ("%Y-%m-%d","%m/%d/%Y","%m/%d/%y","%Y/%m/%d"):
        try: return datetime.strptime(s, fmt)
        except ValueError: continue
    raise ValueError(f"Unparseable DOB: {repr(s)}")


def _parse_eff(s: str) -> datetime:
    for fmt in ("%Y-%m-%d","%m/%d/%Y"):
        try: return datetime.strptime((s or "").strip(), fmt)
        except ValueError: continue
    raise ValueError(f"Unparseable effective_date: {repr(s)}")


def _admin_cell(admin: str) -> str:
    a = (admin or "").strip().upper()
    if a == "EBPA": return "EBPA "
    if "HEALTHEZ" in a and "RBP" in a: return "HEALTHEZ (RBP)"
    if a.replace("_","") == "HEALTHEZ": return "HEALTHEZ (RBP)"
    if a in ("CBA BLUE","CBA","CBA_BLUE"): return "CBA BLUE"
    if "CIGNA" in a: return "HEALTHEZ (CIGNA)"
    if "FREEDOM" in a: return "HEALTHEZ (FREEDOM)"
    return admin


def _area_cell(area: str) -> str:
    m = {"birmingham":"Birmingham","huntsville":"Huntsville",
         "montgomery":"Montgomery","alabama other area":"Alabama Other Area",
         "al other":"Alabama Other Area","out-of-state":"Out-of-State",
         "out of state":"Out-of-State","oos":"Out-of-State"}
    return m.get((area or "").strip().lower(), area or "Birmingham")


def _date_serial(d: datetime) -> int:
    """Excel/LO date serial: days since 1899-12-30."""
    return (d - datetime(1899, 12, 30)).days


# ══ PATH A — UNO ════════════════════════════════════════════════════════════
def _has_uno() -> bool:
    try:
        import uno  # noqa: F401
        return True
    except ImportError:
        return False


def _start_soffice(profile: Path, port: int) -> subprocess.Popen:
    cmd = [
        "soffice","--headless","--norestore","--nologo","--nocrashreport",
        "--nodefault","--nofirststartwizard","--nolockcheck",
        f"--accept=socket,host=127.0.0.1,port={port};urp;StarOffice.ServiceManager",
        f"-env:UserInstallation=file://{profile}",
    ]
    return subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)


def _connect_uno(port: int, timeout: float = 60.0):
    import uno  # type: ignore
    deadline = time.time() + timeout
    last: Exception | None = None
    tries = 0
    while time.time() < deadline:
        try:
            ctx = uno.getComponentContext()
            res = ctx.ServiceManager.createInstanceWithContext(
                "com.sun.star.bridge.UnoUrlResolver", ctx)
            sctx = res.resolve(
                f"uno:socket,host=127.0.0.1,port={port};urp;StarOffice.ComponentContext")
            desk = sctx.ServiceManager.createInstanceWithContext(
                "com.sun.star.frame.Desktop", sctx)
            return desk, sctx
        except Exception as e:
            last = e
            tries += 1
            time.sleep(0.4)
    raise RuntimeError(f"UNO connect timed out after {timeout}s ({tries} tries): {last}")


def _prop(name: str, value):
    from com.sun.star.beans import PropertyValue  # type: ignore
    p = PropertyValue(); p.Name = name; p.Value = value; return p


def run_path_uno(template: Path, work: Path,
                 group: str, eff: datetime, area: str, admin: str,
                 census: list[dict]) -> dict:
    """Full UNO pipeline: open xlsm, inject, calculate, read."""
    import uno  # noqa: F401 (already checked)

    token = uuid.uuid4().hex[:10]
    filled_xlsm = work / f"filled_{token}.xlsm"
    shutil.copy(template, filled_xlsm)
    _log(f"[uno] copied template -> {filled_xlsm.name} ({filled_xlsm.stat().st_size} bytes)")

    profile = work / f"up_{token}"
    profile.mkdir(parents=True, exist_ok=True)
    port = 2002 + (os.getpid() % 997) + (int(time.time()) % 100)
    _log(f"[uno] spawning soffice port={port}")
    proc = _start_soffice(profile, port)

    try:
        t0 = time.time()
        desktop, _ctx = _connect_uno(port, timeout=60.0)
        _log(f"[uno] connected in {time.time()-t0:.1f}s")

        url = "file://" + str(filled_xlsm.resolve())
        t0 = time.time()
        load_props = (_prop("Hidden", True), _prop("ReadOnly", False),
                      _prop("MacroExecutionMode", 4), _prop("UpdateDocMode", 1))
        doc = desktop.loadComponentFromURL(url, "_blank", 0, load_props)
        if doc is None:
            raise RuntimeError("loadComponentFromURL returned None")
        _log(f"[uno] loaded doc in {time.time()-t0:.1f}s")

        sheets = doc.Sheets
        inp = sheets.getByName("Inputs")
        cen = sheets.getByName("Census")

        # Write Inputs (4 cells)
        inp.getCellByPosition(*INP_B4_0).setString(group)
        inp.getCellByPosition(*INP_B5_0).setValue(_date_serial(eff))
        inp.getCellByPosition(*INP_B7_0).setString(_area_cell(area))
        inp.getCellByPosition(*INP_E7_0).setString(_admin_cell(admin))
        _log(f"[uno] Inputs: B4={group!r} B5_serial={_date_serial(eff)} B7={_area_cell(area)!r} E7={_admin_cell(admin)!r}")

        # BULK CLEAR census rows via a cell range (way faster than per-cell)
        # Clear col C..K (2..10) in rows 17..316 (0-indexed)
        t0 = time.time()
        clear_range = cen.getCellRangeByPosition(
            CEN_COL_SSN, CENSUS_FIRST_ROW_0,
            CEN_COL_COST, CENSUS_FIRST_ROW_0 + CENSUS_MAX_ROWS - 1)
        # Flag 31 = VALUE|DATETIME|STRING|FORMULA|ANNOTATION|OBJECTS — clear everything
        clear_range.clearContents(31)
        _log(f"[uno] cleared census range in {time.time()-t0:.1f}s")

        # BULK WRITE census via setDataArray (one UNO call total)
        t0 = time.time()
        n = min(len(census), CENSUS_MAX_ROWS)
        # Build 2D array for cols SSN..COST (C..K = 9 cols)
        # setDataArray writes strings OR numbers; dates we'll handle separately
        data_rows = []
        dob_updates = []  # list of (row_idx_in_census, dob_datetime)
        for i in range(n):
            m = census[i]
            row = [
                f"SSN{i+1:05d}",                                                 # C
                _norm_rel(m.get("relationship","")),                             # D
                str(m.get("first_name") or m.get("firstName") or f"M{i+1}"),     # E
                str(m.get("last_name")  or m.get("lastName")  or "Doe"),         # F
                "",  # G (DOB) - written separately as numeric date
                "",  # H (Company)
                "",  # I (Plan)
                "",  # J (Coverage)
                "",  # K (Cost)
            ]
            data_rows.append(tuple(row))
            dob = m.get("dob") or m.get("date_of_birth") or m.get("dateOfBirth")
            if dob:
                try:
                    dobd = dob if isinstance(dob, datetime) else _parse_dob(str(dob))
                    dob_updates.append((i, dobd))
                except Exception as e:
                    _log(f"[uno] skip bad DOB row {i}: {e}")
        if n > 0:
            write_range = cen.getCellRangeByPosition(
                CEN_COL_SSN, CENSUS_FIRST_ROW_0,
                CEN_COL_COST, CENSUS_FIRST_ROW_0 + n - 1)
            write_range.setDataArray(tuple(data_rows))
            _log(f"[uno] wrote {n} census rows in {time.time()-t0:.1f}s")

        # Write DOBs as numeric date serials (set per cell since they're numbers)
        for i, dobd in dob_updates:
            cen.getCellByPosition(CEN_COL_DOB, CENSUS_FIRST_ROW_0 + i).setValue(_date_serial(dobd))

        # FORCE FULL RECALC
        t0 = time.time()
        doc.calculateAll()
        _log(f"[uno] calculateAll() in {time.time()-t0:.1f}s")

        # Extract rates via UNO
        hdv = sheets.getByName("3. Rate Sheet - HDV")
        rates = {}
        for r0 in HDV_PLAN_ROWS_0:
            plan = hdv.getCellByPosition(0, r0).getString().strip()
            if not plan:
                continue
            def gv(c: int) -> float | None:
                cell = hdv.getCellByPosition(c, r0)
                try: t = int(cell.getType())
                except Exception: t = 0
                v = cell.getValue()
                if t in (1, 3):  # VALUE or FORMULA
                    if v == v:
                        return round(float(v), 2)
                return None
            rates[plan] = {"EE": gv(1), "EC": gv(2), "ES": gv(3), "EF": gv(4)}

        # Diagnostics
        diag = {"path": "uno"}
        try:
            diag["inputs.B4"] = inp.getCellByPosition(*INP_B4_0).getString()
            diag["inputs.B5.serial"] = inp.getCellByPosition(*INP_B5_0).getValue()
            diag["inputs.B7"] = inp.getCellByPosition(*INP_B7_0).getString()
            diag["inputs.E7"] = inp.getCellByPosition(*INP_E7_0).getString()
        except Exception as e:
            diag["inputs.err"] = str(e)
        try:
            rsa = sheets.getByName("Rate Summary All")
            diag["rsa.J131"] = rsa.getCellByPosition(9, 130).getValue()
            diag["rsa.C131"] = rsa.getCellByPosition(2, 130).getValue()
            diag["rsa.J128"] = rsa.getCellByPosition(9, 127).getValue()
        except Exception as e:
            diag["rsa.err"] = str(e)
        diag["census.requested"] = len(census)
        try:
            pop = 0
            for i in range(100):
                if cen.getCellByPosition(CEN_COL_REL, CENSUS_FIRST_ROW_0 + i).getString().strip():
                    pop += 1
            diag["census.populated_rows"] = pop
        except Exception as e:
            diag["census.err"] = str(e)

        # Save recalc'd xlsx for audit (non-fatal on error)
        try:
            out_xlsx = work / f"recalc_{token}.xlsx"
            doc.storeToURL("file://" + str(out_xlsx.resolve()),
                           (_prop("FilterName","Calc Office Open XML"), _prop("Overwrite", True)))
            _log(f"[uno] saved recalc'd xlsx")
        except Exception as e:
            _log(f"[uno] storeToURL failed non-fatally: {e}")

        doc.close(True)
        _log(f"[uno] plan_rates count={len(rates)} nonzero_EE={sum(1 for r in rates.values() if r.get('EE'))}")
        return {"plan_rates": rates, "diagnostics": diag}
    finally:
        try:
            proc.terminate(); proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            proc.kill()
        except Exception:
            pass


# ══ PATH B — openpyxl + soffice --convert-to xlsx ═══════════════════════════
def run_path_cli(template: Path, work: Path,
                 group: str, eff: datetime, area: str, admin: str,
                 census: list[dict]) -> dict:
    """Fallback: openpyxl inject + soffice --convert-to xlsx (with recalc flag)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl not installed and python-uno also missing — both engines unavailable")

    token = uuid.uuid4().hex[:10]
    filled_xlsm = work / f"filled_{token}.xlsm"
    shutil.copy(template, filled_xlsm)
    _log(f"[cli] copied template")

    t0 = time.time()
    wb = load_workbook(filled_xlsm, keep_vba=True, data_only=False, keep_links=False)
    _log(f"[cli] openpyxl load in {time.time()-t0:.1f}s")

    inp = wb["Inputs"]
    inp["B4"] = group
    inp["B5"] = eff
    inp["B7"] = _area_cell(area)
    inp["E7"] = _admin_cell(admin)

    cen = wb["Census"]
    for r in range(CENSUS_FIRST_ROW_1, CENSUS_FIRST_ROW_1 + CENSUS_MAX_ROWS):
        for c in range(CEN_COL_SSN+1, CEN_COL_COST+2):   # 1-indexed
            cen.cell(r, c).value = None
    for i, m in enumerate(census):
        if i >= CENSUS_MAX_ROWS: break
        r = CENSUS_FIRST_ROW_1 + i
        cen.cell(r, CEN_COL_SSN+1).value = f"SSN{i+1:05d}"
        cen.cell(r, CEN_COL_REL+1).value = _norm_rel(m.get("relationship",""))
        cen.cell(r, CEN_COL_FN+1).value  = str(m.get("first_name") or m.get("firstName") or f"M{i+1}")
        cen.cell(r, CEN_COL_LN+1).value  = str(m.get("last_name")  or m.get("lastName")  or "Doe")
        dob = m.get("dob") or m.get("date_of_birth") or m.get("dateOfBirth")
        if dob:
            try: cen.cell(r, CEN_COL_DOB+1).value = dob if isinstance(dob, datetime) else _parse_dob(str(dob))
            except Exception: pass

    t0 = time.time()
    wb.save(filled_xlsm)
    _log(f"[cli] openpyxl save in {time.time()-t0:.1f}s")

    # Build user profile with OOXML recalc-on-load flag
    profile = work / f"cliup_{token}"
    user = profile / "user"
    user.mkdir(parents=True, exist_ok=True)
    (user / "registrymodifications.xcu").write_text(
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<oor:items xmlns:oor="http://openoffice.org/2001/registry">\n'
        ' <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>1</value></prop></item>\n'
        ' <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="ODFRecalcMode" oor:op="fuse"><value>1</value></prop></item>\n'
        '</oor:items>\n')
    cmd = ["soffice","--headless","--calc","--nologo","--nofirststartwizard",
           "--nocrashreport","--nodefault",
           f"-env:UserInstallation=file://{profile}",
           "--convert-to","xlsx","--outdir",str(work),str(filled_xlsm)]
    t0 = time.time()
    r = subprocess.run(cmd, capture_output=True, timeout=150)
    _log(f"[cli] soffice convert rc={r.returncode} in {time.time()-t0:.1f}s")
    if r.returncode != 0:
        raise RuntimeError(f"soffice convert rc={r.returncode} stderr={r.stderr.decode(errors='replace')[-300:]}")

    out_xlsx = work / (filled_xlsm.stem + ".xlsx")
    if not out_xlsx.exists():
        cands = [f for f in work.iterdir() if f.suffix.lower()==".xlsx" and f.stem==filled_xlsm.stem]
        if not cands: raise RuntimeError("no xlsx produced")
        out_xlsx = cands[0]

    t0 = time.time()
    wb2 = load_workbook(out_xlsx, data_only=True, read_only=True, keep_links=False)
    ws = wb2["3. Rate Sheet - HDV"]
    rates = {}
    def n(v):
        if isinstance(v,(int,float)) and v==v: return round(float(v),2)
        return None
    for r0 in HDV_PLAN_ROWS_0:
        row = r0+1
        name = ws.cell(row,1).value
        if not name or not isinstance(name,str) or not name.strip(): continue
        rates[name.strip()] = {"EE": n(ws.cell(row,2).value), "EC": n(ws.cell(row,3).value),
                                "ES": n(ws.cell(row,4).value), "EF": n(ws.cell(row,5).value)}
    _log(f"[cli] extract in {time.time()-t0:.1f}s, {len(rates)} plans")

    diag = {"path":"cli"}
    try:
        inp2 = wb2["Inputs"]
        diag["inputs.B4"] = inp2["B4"].value
        diag["inputs.B5"] = str(inp2["B5"].value)
        diag["inputs.B7"] = inp2["B7"].value
        diag["inputs.E7"] = inp2["E7"].value
    except Exception as e: diag["inputs.err"] = str(e)
    try:
        rsa = wb2["Rate Summary All"]
        diag["rsa.J131"] = rsa["J131"].value
        diag["rsa.C131"] = rsa["C131"].value
        diag["rsa.J128"] = rsa["J128"].value
    except Exception as e: diag["rsa.err"] = str(e)
    return {"plan_rates": rates, "diagnostics": diag}


# ══ dispatcher ══════════════════════════════════════════════════════════════
def run_pipeline(*args, **kwargs) -> dict:
    if _has_uno():
        _log("dispatcher: python-uno available -> path A")
        try:
            return run_path_uno(*args, **kwargs)
        except Exception as e:
            _log(f"path A (uno) FAILED: {type(e).__name__}: {e}")
            _log("traceback:\n" + traceback.format_exc())
            _log("falling back to path B (cli)")
            return run_path_cli(*args, **kwargs)
    else:
        _log("dispatcher: python-uno NOT available -> path B only")
        return run_path_cli(*args, **kwargs)


def summary(census: list[dict], eff: datetime) -> dict:
    def age_at(s):
        try: d = _parse_dob(s)
        except Exception: return None
        y = eff.year - d.year
        if (eff.month, eff.day) < (d.month, d.day): y -= 1
        return max(0, y)
    ees = [m for m in census if _norm_rel(m.get("relationship", "")) == "Employee"]
    ages = [age_at(str(m.get("dob") or m.get("date_of_birth") or m.get("dateOfBirth") or "")) for m in census]
    ages = [a for a in ages if a is not None]
    return {"n_members": len(census), "n_employees": len(ees),
            "avg_age": round(sum(ages)/len(ages), 1) if ages else 0.0}


def main() -> None:
    try: req = json.loads(sys.stdin.read() or "{}")
    except Exception as e: _fail(f"bad stdin json: {e}")

    _log(f"engine start: py={sys.version.split()[0]} uno={'YES' if _has_uno() else 'NO'}")
    tpl = req.get("template_path")
    if not tpl or not Path(tpl).exists(): _fail(f"template missing: {tpl}")
    work = Path(req.get("work_dir") or tempfile.mkdtemp(prefix="xrate_"))
    work.mkdir(parents=True, exist_ok=True)

    try: eff = _parse_eff(str(req.get("effective_date", "")))
    except Exception as e: _fail(f"bad eff date: {e}")

    census = req.get("census") or []
    if not isinstance(census, list) or not census: _fail("census empty")

    group = str(req.get("group") or "Kennion Proposal")
    area = str(req.get("rating_area") or "Birmingham")
    admin = str(req.get("admin") or "EBPA")

    t0 = time.time()
    try:
        result = run_pipeline(Path(tpl), work, group, eff, area, admin, census)
    except Exception as e:
        _log("pipeline FATAL:\n" + traceback.format_exc())
        _fail(f"pipeline: {type(e).__name__}: {e}")
    dt = time.time() - t0
    _log(f"pipeline done in {dt:.1f}s path={result.get('diagnostics',{}).get('path','?')}")

    sys.stdout.write(json.dumps({
        "engine_version": "xlsm-2.1",
        "group": group, "effective_date": eff.strftime("%Y-%m-%d"),
        "rating_area": _area_cell(area), "admin": admin,
        **summary(census, eff),
        "plan_rates": result["plan_rates"],
        "diagnostics": result.get("diagnostics", {}),
        "timings_sec": {"total": round(dt, 2)},
    }) + "\n")


if __name__ == "__main__":
    main()

"""Sync factor tables from the actuary's .xlsm into JSON consumed by the engine.

Run whenever the actuary updates `Kennion Actuarial Rater.xlsm`:

    python3 sync_rater.py --xlsm "../Kennion Actuarial Rater.xlsm" --out factor_tables_v2.json

Then point the engine at the new file:

    python3 kennion_rate_engine.py ... --tables factor_tables_v2.json

Uses openpyxl read-only mode + data_only so the ~14 MB workbook loads without
evaluating VBA or formulas — we only read the cached numeric values that Excel
last saved.
"""
from __future__ import annotations

import argparse
import datetime
import hashlib
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    sys.stderr.write("openpyxl is required. pip install openpyxl\n")
    sys.exit(1)


REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSM = REPO_ROOT / "Kennion Actuarial Rater.xlsm"
DEFAULT_OUT = REPO_ROOT / "server" / "factor-tables.json"

AGE_TAB = "Age Rating Factors"
AREA_TAB = "Area Rating Factors"
PLAN_TAB = "Plan Base Rates"

EXPECTED_AREAS = ["Birmingham", "Huntsville", "Montgomery",
                  "Alabama Other Area", "Out-of-State"]

DEFAULT_TIERS = {"EE": 1.00, "ECH": 1.85, "ESP": 2.00, "FAM": 2.85}


def _file_sha(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()[:16]


def extract_tables(xlsm: Path) -> dict:
    wb = load_workbook(xlsm, data_only=True, read_only=True, keep_links=False)
    tables: dict = {
        "version": "1.0",
        "source_file": xlsm.name,
        "source_sha256": _file_sha(xlsm),
        "synced_at": datetime.datetime.utcnow().isoformat() + "Z",
    }

    # Age factors: Age col B (2), Factor col C (3), data from row 5.
    if AGE_TAB not in wb.sheetnames:
        raise RuntimeError(f"Sheet {AGE_TAB!r} not found")
    ws = wb[AGE_TAB]
    ages = {}
    for row in ws.iter_rows(min_row=5, max_row=200, values_only=True):
        if not row or len(row) < 3:
            continue
        a, f = row[1], row[2]
        if a is None:
            break
        try:
            ages[int(a)] = float(f)
        except (TypeError, ValueError):
            continue
    tables["age_factors"] = ages

    # Area factors: Area col A (1), Factor col B (2), data rows 4-8.
    area_factors: dict = {}
    if AREA_TAB in wb.sheetnames:
        ws = wb[AREA_TAB]
        for row in ws.iter_rows(min_row=4, max_row=20, values_only=True):
            if not row or len(row) < 2:
                continue
            a, f = row[0], row[1]
            if not isinstance(a, str) or not isinstance(f, (int, float)):
                continue
            area_factors[a.strip()] = float(f)
    tables["area_factors"] = area_factors or {a: None for a in EXPECTED_AREAS}

    # Plan Base Rates tab contains trend, stoploss, expenses, and the plan grid.
    if PLAN_TAB not in wb.sheetnames:
        raise RuntimeError(f"Sheet {PLAN_TAB!r} not found")
    ws = wb[PLAN_TAB]
    rows = list(ws.iter_rows(min_row=1, max_row=90, values_only=True))

    def cell(r, c):
        if r - 1 >= len(rows):
            return None
        row = rows[r - 1]
        return row[c - 1] if c - 1 < len(row) else None

    tables["tier_factors_default"] = dict(DEFAULT_TIERS)
    tables["trend_rate"] = 0.07
    tables["phic_stoploss"] = 1.20
    tables["fixed_expense_pct"] = 0.2905
    tables["expense_pepm"] = {
        "EBPA": None, "HEALTHEZ": None,
        "Virtual_RBP": None, "Virtual_RBP_HEALTHEZ": None,
    }

    for r in range(1, 30):
        label = cell(r, 1)
        if isinstance(label, str):
            L = label.strip().lower()
            if L == "trend rate":
                v = cell(r, 2)
                if isinstance(v, (int, float)):
                    tables["trend_rate"] = float(v)
            elif "phic" in L:
                v = cell(r, 3) or cell(r, 2)
                if isinstance(v, (int, float)):
                    tables["phic_stoploss"] = float(v)
        label2 = cell(r, 2)
        if isinstance(label2, str):
            if label2.strip() == "Fixed Expenses % of Premium":
                v = cell(r, 3)
                if isinstance(v, (int, float)):
                    tables["fixed_expense_pct"] = float(v)
            elif label2.strip().startswith("Total Expenses PEPM"):
                tables["expense_pepm"]["EBPA"] = cell(r, 3)
                tables["expense_pepm"]["HEALTHEZ"] = cell(r, 4)
                tables["expense_pepm"]["Virtual_RBP"] = cell(r, 5)
                tables["expense_pepm"]["Virtual_RBP_HEALTHEZ"] = cell(r, 6)

    # Locate "New Business Plan" header, then walk down reading plan rows.
    # Col F (6) = claims + expenses + margin normalized PMPM = "total_margin".
    # Col J (10) = the 3:1 base rate equivalent.
    plans_header_row = None
    for r in range(1, 30):
        if cell(r, 1) == "New Business Plan":
            plans_header_row = r
            break
    if plans_header_row is None:
        raise RuntimeError("Could not locate 'New Business Plan' header row")

    plans: dict = {}
    for r in range(plans_header_row + 1, 90):
        name = cell(r, 1)
        if not isinstance(name, str) or not name.strip():
            break
        total_margin = cell(r, 6)
        base_3to1 = cell(r, 10)
        if isinstance(total_margin, (int, float)) and total_margin > 0:
            plans[name.strip()] = {
                "total_margin": float(total_margin),
                "base_3to1": float(base_3to1) if isinstance(base_3to1, (int, float)) else None,
                "claims_unadjusted": float(cell(r, 2)) if isinstance(cell(r, 2), (int, float)) else None,
                "agg_sl_adj": float(cell(r, 5)) if isinstance(cell(r, 5), (int, float)) else None,
            }
    tables["plan_base_pmpm_6to1"] = plans

    return tables


def write_json(tables: dict, out_path: Path) -> None:
    tables_out = {
        "version": tables.get("version", "1.0"),
        "source_file": tables["source_file"],
        "source_sha256": tables["source_sha256"],
        "synced_at": tables["synced_at"],
        "age_factors": {str(k): v for k, v in sorted(tables["age_factors"].items())},
        "area_factors": {k: tables["area_factors"].get(k) for k in EXPECTED_AREAS},
        "tier_factors_default": tables["tier_factors_default"],
        "plan_base_pmpm_6to1": {k: v for k, v in sorted(tables["plan_base_pmpm_6to1"].items())},
        "trend_rate": tables["trend_rate"],
        "phic_stoploss": tables["phic_stoploss"],
        "fixed_expense_pct": tables["fixed_expense_pct"],
        "expense_pepm": tables["expense_pepm"],
    }
    out_path.write_text(json.dumps(tables_out, indent=2, default=str))


def main() -> int:
    ap = argparse.ArgumentParser(description="Sync factor tables from .xlsm to JSON.")
    ap.add_argument("--xlsm", default=str(DEFAULT_XLSM))
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    args = ap.parse_args()

    xlsm = Path(args.xlsm)
    if not xlsm.exists():
        sys.stderr.write(f"xlsm not found: {xlsm}\n")
        return 2

    tables = extract_tables(xlsm)

    sys.stderr.write(
        f"[sync_rater] ages={len(tables['age_factors'])}  "
        f"areas={sum(1 for v in tables['area_factors'].values() if v is not None)}/5  "
        f"plans={len(tables['plan_base_pmpm_6to1'])}  "
        f"trend={tables['trend_rate']:.3f}\n"
    )

    write_json(tables, Path(args.out))
    sys.stderr.write(f"[sync_rater] wrote {args.out}\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
